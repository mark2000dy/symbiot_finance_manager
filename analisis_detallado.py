#!/usr/bin/env python3
"""
Análisis detallado de pagos de alumnos
Columnas R-AU (julio 2023 - diciembre 2025)
"""

from openpyxl import load_workbook
import datetime

wb = load_workbook('Gastos Socios Symbiot.xlsx', data_only=True)

print("="*80)
print("ANÁLISIS DETALLADO - INGRESOS ROCKSTARSKULL")
print("="*80)

ws = wb['Ingresos RockstarSkull']

# Leer encabezados
print("\n1. ESTRUCTURA DE LA HOJA:")
print("-" * 80)
print(f"Total filas: {ws.max_row}")
print(f"Total columnas: {ws.max_column}")

# Mostrar todos los encabezados
headers = {}
print("\nEncabezados (columnas A-Q):")
for col in range(1, 18):
    value = ws.cell(1, col).value
    if value:
        col_letter = chr(64 + col) if col <= 26 else f"Col{col}"
        headers[col] = value
        print(f"  {col_letter:3s} (col {col:2d}): {value}")

print("\nEncabezados de PAGOS (columnas R-AU = julio 2023 a diciembre 2025):")
payment_columns = {}
for col in range(18, min(48, ws.max_column + 1)):  # R=18 hasta AU=47
    value = ws.cell(1, col).value
    if value:
        col_letter = chr(64 + col) if col <= 26 else f"Col{col}"
        payment_columns[col] = value
        print(f"  {col_letter:5s} (col {col:2d}): {value}")

# Analizar datos de alumnos
print("\n\n2. DATOS DE ALUMNOS:")
print("-" * 80)

alumnos = []
for row in range(2, ws.max_row + 1):
    alumno_data = {
        'row': row,
        'num_alumno': ws.cell(row, 1).value,  # Col A
        'nombre': ws.cell(row, 2).value,      # Col B
        'fecha_inscripcion': ws.cell(row, 5).value,  # Col E
        'fecha_pago': ws.cell(row, 6).value,  # Col F
        'pagos': {}
    }

    # Leer pagos mensuales
    for col, mes in payment_columns.items():
        value = ws.cell(row, col).value
        if value is not None and value != 0:
            try:
                alumno_data['pagos'][mes] = float(value)
            except:
                alumno_data['pagos'][mes] = value

    # Solo agregar si tiene nombre y al menos un pago
    if alumno_data['nombre'] and alumno_data['pagos']:
        alumnos.append(alumno_data)

print(f"Total alumnos con pagos registrados: {len(alumnos)}")

# Mostrar muestra de primeros 5 alumnos
print("\nMuestra de primeros 5 alumnos:")
for alumno in alumnos[:5]:
    print(f"\n  Alumno: {alumno['nombre']}")
    print(f"    Num: {alumno['num_alumno']}")
    print(f"    Fecha inscripción: {alumno['fecha_inscripcion']}")
    print(f"    Fecha pago: {alumno['fecha_pago']}")
    print(f"    Total de meses con pago: {len(alumno['pagos'])}")
    # Mostrar primeros 3 pagos
    for mes, monto in list(alumno['pagos'].items())[:3]:
        print(f"      {mes}: ${monto:,.2f}")

# Calcular totales por mes
print("\n\n3. TOTALES POR MES (de Excel):")
print("-" * 80)

monthly_totals = {}
monthly_counts = {}

for alumno in alumnos:
    for mes, monto in alumno['pagos'].items():
        if isinstance(monto, (int, float)):
            if mes not in monthly_totals:
                monthly_totals[mes] = 0
                monthly_counts[mes] = 0
            monthly_totals[mes] += monto
            monthly_counts[mes] += 1

for mes in sorted(monthly_totals.keys()):
    print(f"{mes:20s}: {monthly_counts[mes]:3d} pagos | Total: ${monthly_totals[mes]:12,.2f}")

print(f"\nGRAND TOTAL: ${sum(monthly_totals.values()):,.2f}")
print(f"Total de pagos: {sum(monthly_counts.values())}")

# Analizar Gastos RockstarSkull (números negativos)
print("\n\n4. ANÁLISIS GASTOS ROCKSTARSKULL:")
print("-" * 80)

ws_gastos = wb['Gastos RockstarSkull']

gastos_negativos = []
gastos_positivos = []

for row in range(2, ws_gastos.max_row + 1):
    total_cell = ws_gastos.cell(row, 8)  # Columna H (Total)
    total = total_cell.value

    if isinstance(total, (int, float)) and total != 0:
        registro = {
            'row': row,
            'fecha': ws_gastos.cell(row, 1).value,
            'concepto': ws_gastos.cell(row, 2).value,
            'total': total
        }

        if total < 0:
            gastos_negativos.append(registro)
        else:
            gastos_positivos.append(registro)

print(f"Registros con monto NEGATIVO (debería ser ingreso): {len(gastos_negativos)}")
print(f"Registros con monto POSITIVO (gasto normal): {len(gastos_positivos)}")

if gastos_negativos:
    print("\nMuestra de gastos NEGATIVOS (primeros 5):")
    for reg in gastos_negativos[:5]:
        print(f"  Fila {reg['row']}: {reg['fecha']} | {reg['concepto']} | ${reg['total']:,.2f}")

    total_negativos = sum(r['total'] for r in gastos_negativos)
    print(f"\nTotal de gastos negativos: ${total_negativos:,.2f}")
    print(f"Esto debería convertirse a INGRESOS: ${abs(total_negativos):,.2f}")

wb.close()

print("\n" + "="*80)
print("ANÁLISIS COMPLETADO - NO SE MODIFICÓ NADA")
print("="*80)
