#!/usr/bin/env python3
"""
FASE 3: Importar pagos de alumnos desde Excel
Columnas R-AU (julio 2023 - diciembre 2025)
"""

from openpyxl import load_workbook
import mysql.connector
from datetime import datetime
import calendar

# Configuración de BD
DB_CONFIG = {
    'host': 'localhost',
    'database': 'gastos_app_db',
    'user': 'gastos_user',
    'password': 'Gastos2025!'
}

# Mapeo de meses Excel → (año, mes)
# El Excel tiene columnas sin año explícito, pero siguen un orden cronológico
MONTH_MAPPING = {
    'Julio': (2023, 7),      # Col R
    'Agosto': (2023, 8),     # Col S
    'Septiembre': (2023, 9), # Col T
    'Octubre': (2023, 10),   # Col U
    'Noviembre': (2023, 11), # Col V
    'Diciembre': (2023, 12), # Col W
    'Enero': (2024, 1),      # Col X
    'Febrero': (2024, 2),    # Col Y
    'Marzo': (2024, 3),      # Col Z
    'Abril': (2024, 4),      # Col AA (27)
    'Mayo': (2024, 5),       # Col AB (28)
    'Junio': (2024, 6),      # Col AC (29)
    'Julio2': (2024, 7),     # Col AD (30)
    'Agosto2': (2024, 8),    # Col AE (31)
    'Septiembre2': (2024, 9),# Col AF (32)
    'Octubre2': (2024, 10),  # Col AG (33)
    'Noviembre2': (2024, 11),# Col AH (34)
    'Diciembre3': (2024, 12),# Col AI (35)
    'Enero2': (2025, 1),     # Col AJ (36)
    'Febrero2': (2025, 2),   # Col AK (37)
    'Marzo2': (2025, 3),     # Col AL (38)
    'Abril2': (2025, 4),     # Col AM (39)
    'Mayo2': (2025, 5),      # Col AN (40)
    'Junio2': (2025, 6),     # Col AO (41)
    'Julio3': (2025, 7),     # Col AP (42)
    'Agosto3': (2025, 8),    # Col AQ (43)
    'Septiembre3': (2025, 9),# Col AR (44)
    'Octubre3': (2025, 10),  # Col AS (45)
    'Noviembre3': (2025, 11),# Col AT (46)
    'Diciembre2': (2025, 12) # Col AU (47)
}

print("=" * 80)
print("FASE 3: IMPORTAR PAGOS DE ALUMNOS")
print("=" * 80)
print()

# Conectar a BD
try:
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    print("[OK] Conectado a base de datos")
except Exception as e:
    print(f"[ERROR] Error conectando a BD: {e}")
    exit(1)

# Cargar Excel
wb = load_workbook('Gastos Socios Symbiot.xlsx', data_only=True)
ws = wb['Ingresos RockstarSkull']

print(f"[OK] Excel cargado: {ws.max_row} filas\n")

# Leer encabezados de pagos (columnas R-AU)
payment_columns = {}
for col in range(18, 48):  # R=18 hasta AU=47
    header = ws.cell(1, col).value
    if header and header in MONTH_MAPPING:
        payment_columns[col] = header

print(f"Columnas de pago encontradas: {len(payment_columns)}")
print()

# Estadísticas
stats = {
    'alumnos_procesados': 0,
    'pagos_insertados': 0,
    'pagos_duplicados': 0,
    'pagos_error': 0,
    'total_monto': 0
}

# Procesar cada alumno
print("Procesando alumnos...")
print("-" * 80)

for row_num in range(2, ws.max_row + 1):
    # Leer datos del alumno
    num_alumno = ws.cell(row_num, 1).value  # Col A
    nombre_alumno = ws.cell(row_num, 2).value  # Col B
    fecha_inscripcion = ws.cell(row_num, 5).value  # Col E
    forma_pago = ws.cell(row_num, 11).value  # Col K

    # Validar datos básicos
    if not nombre_alumno:
        continue

    # Obtener día de inscripción (día de pago mensual)
    dia_pago = 1
    if isinstance(fecha_inscripcion, datetime):
        dia_pago = fecha_inscripcion.day

    # Procesar pagos mensuales
    pagos_alumno = 0

    for col_num, mes_header in payment_columns.items():
        monto = ws.cell(row_num, col_num).value

        # Validar monto
        if not isinstance(monto, (int, float)) or monto <= 0:
            continue

        # Obtener año y mes
        year, month = MONTH_MAPPING[mes_header]

        # Ajustar día si excede el último día del mes
        max_day = calendar.monthrange(year, month)[1]
        dia_final = min(dia_pago, max_day)

        # Crear fecha de pago
        fecha_pago = f"{year:04d}-{month:02d}-{dia_final:02d}"

        # Crear concepto
        concepto = f"Mensualidad - {nombre_alumno}"

        # Verificar si ya existe (evitar duplicados)
        check_query = """
            SELECT id FROM transacciones
            WHERE concepto = %s
            AND fecha = %s
            AND total = %s
            AND empresa_id = 2
        """
        cursor.execute(check_query, (concepto, fecha_pago, monto))
        existe = cursor.fetchone()

        if existe:
            stats['pagos_duplicados'] += 1
            continue

        # Insertar transacción
        # Nota: 'total' es columna calculada (cantidad * precio_unitario)
        # NO existe columna 'categoria'
        try:
            insert_query = """
                INSERT INTO transacciones
                (empresa_id, tipo, concepto, fecha, forma_pago, cantidad, precio_unitario, socio, created_by, created_at)
                VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """

            cursor.execute(insert_query, (
                1,  # empresa_id: Rockstar Skull (1 = Rockstar Skull, 2 = Symbiot Technologies)
                'I',  # tipo: Ingreso
                concepto,  # concepto
                fecha_pago,  # fecha
                forma_pago or 'Transferencia',  # forma_pago
                1.00,  # cantidad
                monto,  # precio_unitario (total = cantidad * precio_unitario)
                'Sistema',  # socio
                1  # created_by: admin
            ))

            stats['pagos_insertados'] += 1
            stats['total_monto'] += monto
            pagos_alumno += 1

        except Exception as e:
            stats['pagos_error'] += 1
            print(f"[ERROR] Error insertando pago: {e}")

    if pagos_alumno > 0:
        stats['alumnos_procesados'] += 1
        num_str = str(num_alumno) if num_alumno else "N/A"
        print(f"  {num_str:5s} | {nombre_alumno[:40]:40s} | {pagos_alumno:2d} pagos insertados")

# Commit de cambios
conn.commit()

print()
print("=" * 80)
print("[OK] FASE 3 COMPLETADA")
print("=" * 80)
print(f"Alumnos procesados:     {stats['alumnos_procesados']}")
print(f"Pagos insertados:       {stats['pagos_insertados']}")
print(f"Pagos duplicados:       {stats['pagos_duplicados']}")
print(f"Pagos con error:        {stats['pagos_error']}")
print(f"Monto total insertado:  ${stats['total_monto']:,.2f}")
print()

# Cerrar conexiones
cursor.close()
conn.close()
wb.close()

print("[OK] Importacion completada")
