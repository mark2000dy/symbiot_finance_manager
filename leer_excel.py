#!/usr/bin/env python3
"""
Script para leer y analizar Gastos Socios Symbiot.xlsx
Solo análisis, NO modificaciones
"""

try:
    from openpyxl import load_workbook
    import json

    # Cargar el archivo Excel
    wb = load_workbook('Gastos Socios Symbiot.xlsx', data_only=True)

    print("="*60)
    print("ANÁLISIS DE EXCEL: Gastos Socios Symbiot.xlsx")
    print("="*60)
    print(f"\nPestañas encontradas: {wb.sheetnames}\n")

    # Analizar cada pestaña
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"PESTAÑA: {sheet_name}")
        print(f"{'='*60}")
        print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")

        # Mostrar encabezados (primera fila)
        headers = []
        for col in range(1, min(ws.max_column + 1, 50)):  # Primeras 50 columnas
            cell_value = ws.cell(1, col).value
            if cell_value:
                headers.append(f"{chr(64+col) if col <= 26 else 'Col'+str(col)}: {cell_value}")

        if headers:
            print("\nEncabezados:")
            for header in headers[:10]:  # Primeros 10
                print(f"  {header}")
            if len(headers) > 10:
                print(f"  ... y {len(headers)-10} más")

        # Si es una pestaña de datos, calcular totales
        if 'gastos' in sheet_name.lower() or 'ingresos' in sheet_name.lower():
            # Buscar columna de totales
            total_sum = 0
            count = 0

            # Intentar sumar valores numéricos en filas
            for row in range(2, min(ws.max_row + 1, 1000)):  # Limitar a 1000 filas
                for col in range(1, min(ws.max_column + 1, 10)):  # Primeras 10 columnas
                    cell_value = ws.cell(row, col).value
                    if isinstance(cell_value, (int, float)) and cell_value != 0:
                        if col == 4:  # Supongo que columna D es monto
                            total_sum += cell_value
                            count += 1
                            break

            if count > 0:
                print(f"\nRegistros encontrados: {count}")
                print(f"Suma total (estimada): ${total_sum:,.2f}")

    print(f"\n{'='*60}")
    print("ANÁLISIS DE COLUMNAS R-AU (Pagos alumnos julio 2023 - dic 2025)")
    print(f"{'='*60}")

    # Buscar hoja con datos de alumnos
    for sheet_name in wb.sheetnames:
        if 'ingreso' in sheet_name.lower():
            ws = wb[sheet_name]

            # Columnas R=18 hasta AU=47 (julio 2023 - diciembre 2025)
            print(f"\nAnalizando pestaña: {sheet_name}")
            print(f"Columnas de pago (R a AU): {18} a {47}")

            # Leer encabezados de pagos
            payment_headers = []
            for col in range(18, min(48, ws.max_column + 1)):
                header = ws.cell(1, col).value
                if header:
                    payment_headers.append((col, header))

            print(f"\nEncabezados de pagos encontrados: {len(payment_headers)}")
            for col, header in payment_headers[:5]:
                print(f"  Columna {col} ({chr(64+col) if col <= 26 else 'Col'+str(col)}): {header}")

            # Contar pagos por mes
            monthly_totals = {}
            for col, header in payment_headers:
                total = 0
                count = 0
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row, col).value
                    if isinstance(value, (int, float)) and value > 0:
                        total += value
                        count += 1

                if count > 0:
                    monthly_totals[header] = {'count': count, 'total': total}

            print(f"\nResumen de pagos por mes:")
            for month, data in list(monthly_totals.items())[:10]:
                print(f"  {month}: {data['count']} pagos, Total: ${data['total']:,.2f}")

            break

    wb.close()

except ImportError:
    print("ERROR: openpyxl no está instalado")
    print("Instala con: pip install openpyxl")
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
